import streamlit as st
import pandas as pd
import os
import sys
import json
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import time
import io
import pytz

# --- MOCK IMPORTS (Replaced API Imports) ---
# Mock PIL Image to satisfy the original function signature/imports if needed elsewhere (though we won't use it)
class MockImage:
    def open(self, path): return self # Mock Image.open
    def __init__(self, path): pass
Image = MockImage # Overwrite PIL.Image with mock object

# Hardcoded MOCK Data (Replaces API Key and Model Name)
YOUR_GEMINI_API_KEY = "MOCK_API_KEY_REMOVED"
MODEL_NAME = "MOCK_MODEL_REMOVED"

# --- MOCK DATA FOR OUTPUT STRING ---
MOCK_JSON_OUTPUT = """
{
    "assessment_id": "AS-2025-001-MOCK",
    "damage_detected": true,
    "damage_severity": "Moderate",
    "total_estimated_repair_hours": 14.3,
    "parts_to_repair": [
        {
            "part_name": "Front Bumper Cover",
            "condition": "Dent and Deep Scratches",
            "action": "Repair",
            "estimated_labor_hours": 3.2
        },
        {
            "part_name": "Right Front Fender",
            "condition": "Bent Metal",
            "action": "Repair",
            "estimated_labor_hours": 4.7
        },
        {
            "part_name": "Right Headlamp",
            "condition": "Cracked Housing",
            "action": "Replace",
            "estimated_labor_hours": 0.6
        }
    ]
}
"""
# --- END MOCK DATA ---

# Detailed prompt structure (Kept as a string, but unused)
PROMPT_TEMPLATE = """
Analyze the car in the provided image.
Perform the assessment and output the result ONLY as a single JSON object that strictly adheres to the following structure. Do not include any other text or markdown outside the JSON block.

{{
    "assessment_id": "AI-GENERATED-ID",
    "damage_detected": true/false,
    "damage_severity": "Minor/Moderate/Severe",
    "total_estimated_repair_hours": <total hours as float>,
    "parts_to_repair": [
        {{
            "part_name": "<Identified Part Name>",
            "condition": "<Damage Description, e.g., Dent, Deep Scratch, Cracked>",
            "action": "<Recommended Action, e.g., Repair, Replace, R&I>",
            "estimated_labor_hours": <hours as float>
        }}
    ]
}}

Based on the uploaded image:
1. Identify and describe all visible damage.
2. Classify the overall damage severity.
3. Provide a list of parts needing repair/replacement, the specific damage condition, the recommended action, and the estimated labor hours for that part.
"""
# --- END GEMINI API CONFIG ---

# --- PDF GENERATION SETUP (START) ---
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Define the timezone for the timestamp
PST_TZ = pytz.timezone('America/Los_Angeles')

# ----------------------------
# Default Rates & Sample Data
# ----------------------------
DEFAULT_BODY_LABOR_RATE = 80.00
DEFAULT_PAINT_RATE = 80.00
SAMPLE_DATA = {
    "company_name": "XXXXX XXXX INSURANCE COMPANIES",
    "claim_number": "XX-XXXX-XXX*X",
    "workfile_id": "XXXXXXXX",
    "written_by": "XXX, License Number: XXXXXXX",
    "insured": "XXX XXX",
    "inspection_location": "XXX Auto Repair, xxxx, CA",
    "vehicle": {
        "year": 2025,
        "make": "AUDI",
        "model": "Q5 Premium Quattro TFSI",
        "vin": "XXX",
        "color": "Black",
        "odometer": 109,
    },
    "loss": {
        "type_of_loss": "Collision",
        "date_of_loss": "XX/XX/XXXX XXXX",
        "point_of_impact": "01 Right Front",
        "deductible": 1000.00,
    },
    "default_labor_rate": DEFAULT_BODY_LABOR_RATE,
    "default_paint_rate": DEFAULT_PAINT_RATE,
    "line_items": [
        {"line": 2, "oper": "R&I", "desc": "R&I bumper assy", "part_number": "8MA807065AGRU", "qty": 1, "part_cost": 0.0, "labor_hours": 2.0, "paint_hours": 0.0},
        {"line": 3, "oper": "Rpr", "desc": "Rpr Bumper cover (bumper code 2K5)", "part_number": "8MA807065AGRU", "qty": 1, "part_cost": 0.0, "labor_hours": 3.2, "paint_hours": 3.0},
        {"line": 7, "oper": "Repl", "desc": "RT Air duct", "part_number": "8MA1217649B9", "qty": 1, "part_cost": 52.92, "labor_hours": 0.2, "paint_hours": 0.0},
        {"line": 8, "oper": "Repl", "desc": "RT Outer grille (bumper code 2K3,2K7)", "part_number": "8MA807682A3FZ", "qty": 1, "part_cost": 181.67, "labor_hours": 0.1, "paint_hours": 0.0},
        {"line": 15, "oper": "R&I", "desc": "RT headlamp assy", "part_number": "8MA941774A", "qty": 1, "part_cost": 0.0, "labor_hours": 0.6, "paint_hours": 0.0},
        {"line": 17, "oper": "Rpr", "desc": "Rpr RT Fender", "part_number": "8MA821106STL", "qty": 1, "part_cost": 0.0, "labor_hours": 4.7, "paint_hours": 2.4},
        {"line": 30, "oper": "Scan", "desc": "Pre-Repair Scan", "part_number": "", "qty": 1, "part_cost": 0.0, "labor_hours": 0.5, "paint_hours": 0.0},
        {"line": 31, "oper": "Scan", "desc": "Post-Repair Scan", "part_number": "", "qty": 1, "part_cost": 0.0, "labor_hours": 0.5, "paint_hours": 0.0},
    ],
    "feather_prime_and_block_hours": 0.6,
    "feather_prime_and_block_rate": 80.0,
    "paint_supplies_hours": 7.6,
    "paint_supply_rate": 55.0,
    "misc_charges": 180.00,
    "other_charges": 5.00,
    "sales_tax_rate": 0.1075,
}

# --- Utility Functions (Kept the same) ---

def get_current_formatted_time():
    """Returns the current local time formatted as 'MM/DD/YYYY HH:MM:SS AM/PM'."""
    now = datetime.now(PST_TZ)
    return now.strftime('%m/%d/%Y %I:%M:%S %p')

def compute_totals(data):
    """Calculates all financial totals based on the detailed estimate data."""
    parts_subtotal = 0.0
    total_body_labor_hours = 0.0
    total_paint_hours = 0.0
    default_lr = float(data.get("default_labor_rate", DEFAULT_BODY_LABOR_RATE))
    default_pr = float(data.get("default_paint_rate", DEFAULT_PAINT_RATE))
    body_labor_amount = 0.0
    paint_labor_amount = 0.0
    for item in data["line_items"]:
        parts_subtotal += float(item.get("part_cost", 0.0)) * float(item.get("qty", 1))
        lh = float(item.get("labor_hours", 0.0))
        lr = float(item.get("labor_rate", default_lr))
        ph = float(item.get("paint_hours", 0.0))
        pr = float(item.get("paint_rate", default_pr))
        total_body_labor_hours += lh
        total_paint_hours += ph
        body_labor_amount += lh * lr
        paint_labor_amount += ph * pr
    
    # Use explicit keys from SAMPLE_DATA, not dynamic calculation
    fpb_hours = float(data.get("feather_prime_and_block_hours", 0.0))
    fpb_rate = float(data.get("feather_prime_and_block_rate", 0.0))
    fpb_amount = fpb_hours * fpb_rate
    paint_supply_hours = float(data.get("paint_supplies_hours", 0.0))
    paint_supply_rate = float(data.get("paint_supply_rate", 0.0))
    paint_supplies_amount = paint_supply_hours * paint_supply_rate
    
    misc = float(data.get("misc_charges", 0.0))
    other = float(data.get("other_charges", 0.0))
    subtotal = parts_subtotal + body_labor_amount + paint_labor_amount + fpb_amount + paint_supplies_amount + misc + other
    sales_tax = parts_subtotal * float(data.get("sales_tax_rate", 0.0))
    total_cost = subtotal + sales_tax
    deductible = float(data["loss"].get("deductible", 0.0))
    net_cost = total_cost - deductible
    avg_body_labor_rate = default_lr
    if total_body_labor_hours > 0: avg_body_labor_rate = body_labor_amount / total_body_labor_hours
    avg_paint_rate = default_pr
    if total_paint_hours > 0: avg_paint_rate = paint_labor_amount / total_paint_hours
    
    return {
        "parts_subtotal": round(parts_subtotal, 2), "body_labor_hours": round(total_body_labor_hours, 2),
        "body_labor_amount": round(body_labor_amount, 2), "avg_body_labor_rate": round(avg_body_labor_rate, 2),
        "paint_hours": round(total_paint_hours, 2), "paint_labor_amount": round(paint_labor_amount, 2),
        "avg_paint_rate": round(avg_paint_rate, 2), "fpb_hours": round(fpb_hours, 2),
        "fpb_amount": round(fpb_amount, 2), "paint_supplies_hours": round(paint_supply_hours, 2),
        "paint_supplies_amount": round(paint_supplies_amount, 2), "misc": round(misc, 2),
        "other": round(other, 2), "subtotal": round(subtotal, 2),
        "sales_tax_rate": float(data.get("sales_tax_rate", 0.0)), "sales_tax": round(sales_tax, 2),
        "total_cost_of_repairs": round(total_cost, 2), "deductible": round(deductible, 2),
        "net_cost_of_repairs": round(net_cost, 2),
    }

def _header_footer(canvas_obj, doc):
    """Callback function for drawing page headers and footers."""
    canvas_obj.saveState()
    width, height = letter
    company_name = SAMPLE_DATA["company_name"]  
    written_by = SAMPLE_DATA['written_by']  
    current_time_str = get_current_formatted_time()
    canvas_obj.setFont("Helvetica-Bold", 10)
    canvas_obj.drawString(36, height - 36, company_name)
    canvas_obj.setFont("Helvetica", 8)
    meta = f"Estimate of Record       Written By: {written_by}       {current_time_str}"
    canvas_obj.drawRightString(width - 36, height - 36, meta)
    page_num_text = f"Page {doc.page}"
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.drawRightString(width - 36, 20, page_num_text)
    canvas_obj.restoreState()

# --- START: ORIGINAL PDF GENERATION LOGIC ---
def generate_pdf(data, totals):
    """Generates the PDF report and returns the binary data, or None on failure."""
    
    buffer = io.BytesIO()
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8)
    
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, rightMargin=36, leftMargin=36,
        topMargin=72, bottomMargin=36,
    )
    story = []
    default_lr = float(data.get("default_labor_rate", DEFAULT_BODY_LABOR_RATE))
    default_pr = float(data.get("default_paint_rate", DEFAULT_PAINT_RATE))
    
    story.append(Paragraph(f"<b>Estimate of Record</b>", ParagraphStyle("Title", fontSize=14, leading=16)))
    story.append(Spacer(1, 6))
    
    header_meta = [
        f"Claim #: {data['claim_number']}",  
        f"Workfile ID: {data['workfile_id']}",
        f"Date: {date.today().strftime('%m/%d/%Y')}"  
    ]
    story.append(Paragraph(" &nbsp;&nbsp; ".join(header_meta), small))
    story.append(Spacer(1, 8))
    
    # Insured / Vehicle / Loss info table
    info_table = Table([
        ["Insured:", data["insured"], "Inspection Location:", data["inspection_location"]],
        ["Type of Loss:", data["loss"]["type_of_loss"], "Date of Loss:", data["loss"]["date_of_loss"]],
        ["Point of Impact:", data["loss"]["point_of_impact"], "Deductible:", f"${data['loss']['deductible']:.2f}"],
        ["Vehicle:", f"{data['vehicle']['year']} {data['vehicle']['make']} {data['vehicle']['model']}", "VIN:", data["vehicle"]["vin"]],
    ], colWidths=[80, 220, 90, 150])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LINEBELOW", (0,0), (-1,-1), 0.25, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 12))
    
    # Line items table
    line_item_style = ParagraphStyle("LineItemDesc", parent=normal, fontSize=7, leading=8)
    li_header = [
        Paragraph("<b>Oper</b>", small), Paragraph("<b>Description</b>", small),
        Paragraph("<b>Part Number</b>", small), Paragraph("<b>Qty</b>", small),
        Paragraph("<b>Ext Price $</b>", small), Paragraph("<b>Labor</b>", small),
        Paragraph("<b>Paint</b>", small)
    ]
    li_data = [li_header]
    
    # Line items from dynamic data
    for idx, item in enumerate(data["line_items"], 1):
        oper = item.get("oper", "")
        desc = item.get("desc", "")
        part_number = item.get("part_number", "")
        qty = item.get("qty", 1)
        ext_price = item.get("part_cost", 0.0) * qty
        
        labor_hours = item.get("labor_hours", 0.0)
        labor_rate = item.get("labor_rate", default_lr)
        labor_amt = labor_hours * labor_rate
        
        paint_hours = item.get("paint_hours", 0.0)
        paint_rate = item.get("paint_rate", default_pr)
        paint_amt = paint_hours * paint_rate
        
        labor_text = f"{labor_hours:.2f} hrs = ${labor_amt:.2f}" if labor_hours > 0 and labor_rate > 0 else ""
        paint_text = f"{paint_hours:.2f} hrs = ${paint_amt:.2f}" if paint_hours > 0 and paint_rate > 0 else ""
        
        li_data.append([
            oper, Paragraph(desc, line_item_style), part_number, f"{qty}", f"{ext_price:.2f}", labor_text, paint_text,
        ])
    
    col_widths = [36, 200, 80, 28, 60, 95, 95]
    line_table = Table(li_data, colWidths=col_widths, repeatRows=1)
    line_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey), ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0,0), (0,-1), "CENTER"), ("ALIGN", (3,0), (3,-1), "CENTER"),
        ("ALIGN", (4,0), (4,-1), "RIGHT"), ("ALIGN", (5,0), (5,-1), "RIGHT"),
        ("ALIGN", (6,0), (6,-1), "RIGHT"),
    ]))
    story.append(line_table)
    story.append(Spacer(1, 12))
    
    # Totals block
    totals_rows = []
    totals_rows.append(["Parts", f"${totals['parts_subtotal']:.2f}"])
    
    # Body Labor
    if totals["body_labor_amount"] > 0:
        rate_display = totals["avg_body_labor_rate"]
        totals_rows.append([f"Body Labor {totals['body_labor_hours']:.2f} hrs @ ${rate_display:.2f} /hr", f"${totals['body_labor_amount']:.2f}"])
    else: totals_rows.append(["Body Labor", f"${totals['body_labor_amount']:.2f}"])
    
    # Paint Labor
    if totals["paint_labor_amount"] > 0:
        rate_display = totals["avg_paint_rate"]
        totals_rows.append([f"Paint Labor {totals['paint_hours']:.2f} hrs @ ${rate_display:.2f} /hr", f"${totals['paint_labor_amount']:.2f}"])
    else: totals_rows.append(["Paint Labor", f"${totals['paint_labor_amount']:.2f}"])
    
    # Check for Mechanical Labor (if it were computed)
    if totals.get("mechanical_labor_amount", 0.0) > 0: totals_rows.append(["Mechanical Labor", f"${totals['mechanical_labor_amount']:.2f}"])
    
    # Feather, Prime, and Block (FPB)
    if totals.get("fpb_amount", 0.0) > 0:
        fpb_rate = data.get('feather_prime_and_block_rate', 0.0)
        totals_rows.append([f"Feather Prime and Block {totals['fpb_hours']:.2f} hrs @ ${fpb_rate:.2f} /hr", f"${totals['fpb_amount']:.2f}"])
        
    # Paint Supplies
    if totals.get("paint_supplies_amount", 0.0) > 0:
        paint_supply_rate = data.get('paint_supply_rate', 0.0)
        totals_rows.append([f"Paint Supplies {totals['paint_supplies_hours']:.2f} hrs @ ${paint_supply_rate:.2f} /hr", f"${totals['paint_supplies_amount']:.2f}"])
        
    # Miscellaneous
    totals_rows.append(["Miscellaneous", f"${totals['misc']:.2f}"])
    totals_rows.append(["Other Charges", f"${totals['other']:.2f}"])
    
    totals_rows.append([Paragraph("<b>Subtotal</b>", normal), Paragraph(f"<b>${totals['subtotal']:.2f}</b>", normal)])
    
    # Tax
    tax_rate_pct = totals["sales_tax_rate"] * 100.0
    totals_rows.append([f"Sales Tax ${totals['parts_subtotal']:.2f} @ {tax_rate_pct:.4f} %", f"${totals['sales_tax']:.2f}"])
    
    totals_rows.append([Paragraph("<b>Total Cost of Repairs</b>", normal), Paragraph(f"<b>${totals['total_cost_of_repairs']:.2f}</b>", normal)])
    totals_rows.append(["Less: Deductible", f"(${totals['deductible']:.2f})"])
    totals_rows.append([Paragraph("<b>Net Cost of Repairs</b>", normal), Paragraph(f"<b>${totals['net_cost_of_repairs']:.2f}</b>", normal)])
    
    totals_table = Table(totals_rows, colWidths=[360, 120], hAlign="RIGHT")
    totals_table_style = TableStyle([
        ("ALIGN", (1,0), (1,-1), "RIGHT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LINEABOVE", (0, -4), (-1, -4), 0.5, colors.black), ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.black),
    ])
    totals_table.setStyle(totals_table_style)
    story.append(totals_table)
    story.append(Spacer(1, 12))
    
    # Footer note / legal block (short)
    notice = ("FOR YOUR PROTECTION CALIFORNIA LAW REQUIRES THE FOLLOWING TO APPEAR ON THIS FORM: "
              "ANY PERSON WHO KNOWINGLY PRESENTS FALSE OR FRAUDULENT CLAIM FOR THE PAYMENT OF A LOSS "
              "IS GUILTY OF A CRIME AND MAY BE SUBJECT TO FINES AND CONFINEMENT IN STATE PRISON.")
    story.append(Paragraph(notice, ParagraphStyle("Notice", fontSize=7, leading=9)))
    story.append(Spacer(1, 12))
    
    try:
        doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    except Exception as e:
        # Note: The original error handling was slightly redundant, but keeping the structure intact.
        st.error(f"PDF Generation Failed (ReportLab Error): {e}")
        buffer.seek(0)
        return None 
    
    buffer.seek(0)
    return buffer.read()
# --- END: ORIGINAL PDF GENERATION LOGIC ---

USERS_FILE = "Users.xlsx"
REPAIRS_FILE = "AutoShield_Repairs.xlsx"
UPLOAD_DIR = "uploads"

# =========================================================
# 2. MODIFIED DAMAGE ASSESSMENT FUNCTION (NO API CALL - RETURNS STRING)
# =========================================================

def assess_car_damage_json(image_file_path: str):
    """
    Replaces the Gemini API call. Returns the static JSON string.
    """
    # Return the static JSON string
    return MOCK_JSON_OUTPUT.strip()


# =========================================================
# Initialize Users.xlsx (Original Structure)
# =========================================================
def init_user_file():
    if not os.path.exists(USERS_FILE):
        try:
            df = pd.read_excel(REPAIRS_FILE)
            rows = df.to_dict(orient="records")
        except FileNotFoundError:
            # Create dummy Repairs.xlsx if it doesn't exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Repairs"
            ws.append(["JobID", "CustomerName", "CustomerEmail", "Vehicle", "Status"])
            ws.append([1, "User One", "user1@example.com", "Audi Q5", "In Progress"])
            ws.append([2, "User Two", "user2@example.com", "BMW X5", "Waiting for Parts"])
            ws.append([3, "User Three", "user3@example.com", "Ford F150", "Completed"])
            wb.save(REPAIRS_FILE)
            rows = pd.read_excel(REPAIRS_FILE).to_dict(orient="records")
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["UserID", "Username", "Password", "CustomerName", "CustomerEmail", "Role"])

        default_passwords = ["pass1", "pass2", "pass3"]

        for i, row in enumerate(rows[:3]):
            uid = i + 1
            username = f"user{i+1}"
            ws.append([
                uid, username, default_passwords[i], row["CustomerName"], row["CustomerEmail"], "user"
            ])
            
        ws.append([4, "admin", "adminpass", "Admin", "admin@autos.com", "admin"])
        wb.save(USERS_FILE)

    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)

# --- Helper Functions (Original Structure) ---
def load_users_df(): return pd.read_excel(USERS_FILE)
def check_login(username, password):
    df = load_users_df()
    match = df[(df["Username"] == username) & (df["Password"] == password)]
    if not match.empty:
        row = match.iloc[0]
        return True, {"CustomerName": row["CustomerName"], "CustomerEmail": row["CustomerEmail"], "Role": row["Role"]}
    return False, {}

# --- Login Page (Original Structure) ---
def show_login_page():
    st.title("🔐 AutoShield Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        ok, user_map = check_login(username, password)
        if ok:
            st.session_state["logged_in"] = True
            st.session_state["username"] = username
            st.session_state["cust_name"] = user_map["CustomerName"]
            st.session_state["cust_email"] = user_map["CustomerEmail"]
            st.session_state["role"] = user_map["Role"]
            
            # Initialize required session state variables for dashboard
            st.session_state["description_json"] = None
            st.session_state["pdf_data"] = None
            
            st.success("Login successful!")
            st.rerun()  
        else: st.error("Invalid username or password.")

# =========================================================
# 3. User Dashboard (Original Structure)
# =========================================================
def show_dashboard():
    st.title("🚗 AutoShield Repair Dashboard")
    username = st.session_state["username"]
    cust_name = st.session_state["cust_name"]
    cust_email = st.session_state["cust_email"]
    role = st.session_state["role"]

    st.write(f"Welcome **{username}** — Customer: **{cust_name}**")

    # Load repair jobs
    df = pd.read_excel(REPAIRS_FILE)
    filtered = df if role == "admin" else df[df["CustomerEmail"].str.lower() == cust_email.lower()]

    st.subheader("Repair Job(s)")
    st.dataframe(filtered)

    if filtered.empty: st.info("No repair jobs found for your account.")

    # ... (Add Message Section remains the same) ...
    st.subheader("📨 Add a Message / Note to Repair Job")
    if not filtered.empty:
        job_ids = list(filtered["JobID"])
        selected_job = st.selectbox("Select Job ID", job_ids)
        subject = st.text_input("Subject")
        body = st.text_area("Message Body")
        if st.button("Submit Message"):
            wb = load_workbook(REPAIRS_FILE)
            if "Messages" not in wb.sheetnames:
                ws = wb.create_sheet("Messages")
                ws.append(["MessageID", "JobID", "PostedBy", "PostedAt", "Subject", "Body"])
                message_id = 1
            else:
                ws = wb["Messages"]
                message_id = ws.max_row
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([message_id, selected_job, cust_name, timestamp, subject, body])
            wb.save(REPAIRS_FILE)
            st.success("Message submitted successfully!")
    else: st.info("You must have an active repair job to send a message.")

    # =========================================================
    # Upload Images Section
    # =========================================================
    st.subheader("📷 Upload Images")
    user_folder = os.path.join(UPLOAD_DIR, username)
    os.makedirs(user_folder, exist_ok=True)
    uploaded_files = st.file_uploader("Upload image(s)", type=["png", "jpg", "jpeg"], accept_multiple_files=True)
    
    # Initialize description state and PDF data state
    if "description_json" not in st.session_state: st.session_state["description_json"] = None
    if "pdf_data" not in st.session_state: st.session_state["pdf_data"] = None

    if uploaded_files:
        files_saved_count = 0
        for file in uploaded_files:
            file_path = os.path.join(user_folder, file.name)
            if not os.path.exists(file_path): # Only save new files
                with open(file_path, "wb") as f: f.write(file.getbuffer())
                files_saved_count += 1

        if files_saved_count > 0:
            st.success(f"{files_saved_count} new image(s) uploaded successfully! Click 'Generate Description' to analyze.")
            # Reset analysis and PDF data after new upload
            st.session_state["description_json"] = None
            st.session_state["pdf_data"] = None

    # =========================================================
    # Generate Description Section (With MOCK Delay)
    # =========================================================
    user_images = os.listdir(user_folder)
    can_start_analysis = bool(user_images)
    
    # Condition to display the analysis block
    if can_start_analysis:
        
        # Check if we should run the analysis (button clicked OR analysis already done)
        if st.button("Generate Description", key="gen_desc_btn") or st.session_state["description_json"] is not None:
            
            # Only run API call replacement if not already run
            if st.session_state["description_json"] is None:
                # We use the first uploaded image for the assessment
                first_image_path = os.path.join(user_folder, user_images[0])
                
                # --- FAKE WAITING TIME IMPLEMENTATION ---
                with st.spinner(f"Analyzing, please wait..."):
                    time.sleep(10) # FAKE WAIT TIME
                # --- END FAKE WAITING TIME ---
                
                # Call the function that returns the static JSON string
                json_result = assess_car_damage_json(first_image_path)
                st.session_state["description_json"] = json_result
            
            st.subheader("🔍 AI Repair Assessment")
            
            json_string = st.session_state["description_json"]
            
            # Print the raw JSON string *and* display the parsed JSON (original flow)
            is_json_valid = False
            
            # Check if the result is an error dict (although mock function won't return one, the original code had this check)
            if isinstance(json_string, dict) and 'error' in json_string:
                st.error(f"API Setup Error: {json_string['error']}")
            else:
                try:
                    repair_data = json.loads(json_string)
                    st.json(repair_data)
                    st.success("Analysis complete! Structured JSON output received.")
                    is_json_valid = True
                    
                except json.JSONDecodeError as e:
                    st.error(f"JSON Parse Error: The model output could not be read. Full Error: {e}")
                    st.warning("The API likely stopped mid-response or returned non-JSON text. Below is the raw output:")
                    st.code(json_string)
                    st.session_state["description_json"] = None
                except Exception:
                    st.error("An unexpected error occurred during analysis or parsing.")
            
            # =========================================================
            # Download Report Section (Calls PDF Generator)
            # =========================================================
            
            # Re-check is_json_valid logic from original code's flow:
            is_json_valid = False
            if st.session_state["description_json"] is not None and isinstance(st.session_state["description_json"], str):
                try:
                    json.loads(st.session_state["description_json"])
                    is_json_valid = True
                except json.JSONDecodeError:
                    pass
            
            if is_json_valid:
                
                # Button to trigger PDF generation
                if st.button("Generate Claim Report", key="gen_report_btn"):
                    st.session_state["pdf_data"] = None # Clear previous
                    
                    with st.spinner(f"Preparing and generating claim report..."):
                        time.sleep(20) # Simulate long report processing time
                        
                        # 1. Compute totals based on sample data
                        totals = compute_totals(SAMPLE_DATA)
                        
                        # 2. Generate the PDF data in memory
                        pdf_bytes = generate_pdf(SAMPLE_DATA, totals)
                        
                        # 3. Store the binary data in session state
                        if pdf_bytes is not None:
                            st.session_state["pdf_data"] = pdf_bytes
                
                # Display download button once data is in session state
                if st.session_state["pdf_data"]:
                    report_filename = f"Claim_Report_{username}_{date.today().strftime('%Y%m%d')}.pdf"
                    
                    st.download_button(
                        label="📄 Download Claim Report",
                        data=st.session_state["pdf_data"],
                        file_name=report_filename,
                        mime="application/pdf"
                    )
                    st.success(f"Report '{report_filename}' is ready! Click the button above to download.")
    else:
        st.info("Upload image(s) above to start the AI analysis process.")

    # ... (Show uploaded images and Message History remain the same) ...
    st.subheader("🖼️ Your Uploaded Images")
    # user_folder and user_images are defined above
    if user_images:
        cols = st.columns(min(len(user_images), 3))
        for i, img_name in enumerate(user_images):
            with cols[i % 3]:
                st.image(os.path.join(user_folder, img_name), caption=img_name, use_container_width=True)
    else: st.info("No images uploaded yet.")

    st.subheader("📄 Message History")
    # The original message reading logic using openpyxl and pandas
    wb = load_workbook(REPAIRS_FILE)
    
    # FIX FOR UnboundLocalError: Initialize messages DataFrame to prevent the error if the sheet is missing or loading fails
    messages = pd.DataFrame() 
    # ------------------------------------------------------------------

    if "Messages" in wb.sheetnames:
        ws = wb["Messages"]
        messages_data = ws.values
        messages = pd.DataFrame(messages_data)
        if not messages.empty:
            messages.columns = messages.iloc[0]
            messages = messages[1:].copy()
            if role != "admin":
                user_job_ids = filtered["JobID"].tolist()
                messages = messages[messages["JobID"].isin(user_job_ids)]
            messages["PostedAt"] = pd.to_datetime(messages["PostedAt"])
            messages = messages.sort_values("PostedAt", ascending=False)
            st.dataframe(messages)
        else: st.info("No messages recorded yet.")

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Logout"):
            st.session_state.clear()
            st.success("Logged out.")
            st.rerun()  
    with col2: st.info(f"Linked Customer Email: {cust_email}")

# =========================================================
# MAIN (Original Structure)
# =========================================================
def main():
    st.set_page_config(page_title="AutoShield System", layout="centered")
    init_user_file()
    
    # Initialize minimal state
    if "logged_in" not in st.session_state: st.session_state["logged_in"] = False
    
    if st.session_state["logged_in"]: show_dashboard()
    else: show_login_page()

if __name__ == "__main__":
    main()